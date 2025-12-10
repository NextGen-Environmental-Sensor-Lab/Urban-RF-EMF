#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Inventory Aggregation and GPS Enhancement Script
------------------------------------------------

This script updates an existing Inventory Excel workbook with per-file RF exposure
statistics and GPS location metadata extracted from individual measurement files.

Workflow:
1) User selects the Inventory Excel file (output target).
2) User selects multiple source Excel measurement files.

For each source file:
- Parses filename to extract date, borough, environment, and location.
- Reads start and end times from the timestamp column.
- Finds GPS coordinate pairs ("GPS Lat", "GPS Lon"):
    * First valid (non-NaN, not (0.0, 0.0)) → start lat, start lon
    * Last  valid (non-NaN, not (0.0, 0.0)) → end   lat, end   lon
    * First valid at rows where Marker == 'x' → trun lat, turn lon
- Computes statistics (MIN, P25, MEAN, GEOMEAN, MEDIAN, P75, P90, MAX, STDEV)
  for category-grouped RMS values (Broadcast, Downlink, Uplink, WLAN, TDD, Total).
- Appends a fully populated row to Sheet1 of the Inventory file.

Aggregated summaries:
- Data is grouped into Totals, Boroughs, and Environments.
- A formatted Sheet2 is generated with summary tables:
    • Label & Category as row/column headers
    • 4-decimal place numeric formatting
    • Bolded header row and label column
    • Single blank line between tables

Output:
- Updates and saves the Inventory workbook in-place.
- Ensures consistent header structure including new GPS columns.
- Prints progress and any processing warnings to console.
"""

import os, re
from math import exp, log
from datetime import datetime, time as dtime
import numpy as np, pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, numbers

# ---------------- CONFIGURATION ----------------

CATEGORY_MAP = {
    "Broadcast": ["FM Radio (RMS)", "VHF 1, 2, 3 (RMS)", "UHF1 (RMS)", "UHF2 (RMS)", "UHF3 (RMS)"],
    "Downlink":  ["Mobile DL (RMS)", "Mobile DL (RMS).1", "Mobile DL (RMS).2", "Mobile DL (RMS).3", "Mobile DL (RMS).4"],
    "Uplink":    ["Mobile UL (RMS).1","Mobile UL (RMS).2","Mobile UL (RMS).3","Mobile UL (RMS).4","Mobile UL (RMS).5"],
    "WLAN":      ["ISM (RMS)","WLAN (RMS)","WLAN (RMS).1","WLAN (RMS).2","WLAN (RMS).3","WLAN (RMS).4",
                  "WLAN (RMS).5","WLAN (RMS).6","WLAN (RMS).7","WLAN (RMS).8","WLAN (RMS).9","WLAN (RMS).10"],
    "TDD":       ["TDD (RMS)","TDD (RMS).1","TDD (RMS).2","TDD (RMS).3","TDD (RMS).4",
                  "TDD (RMS).5","TDD (RMS).6","TDD (RMS).7","TDD (RMS).8"],
}
CATEGORY_MAP["Total"] = sum(CATEGORY_MAP.values(), [])

CATS = ["Broadcast","Downlink","Uplink","WLAN","TDD","Total"]
STATS = ["MIN","P25","MEAN","GEOMEAN","MEDIAN","P75","P90","MAX","STDEV"]

# NOTE: new GPS columns added:
# ... start time, end time, start lat, start lon, trun lat, turn lon, end lat, end lon, N
INVENTORY_HEADER = (
    ["date","borough","location","environment","note",
     "start time","end time",
     "start lat","start lon",
     "trun lat","turn lon",
     "end lat","end lon",
     "N"]
    + [f"{c} {s}" for c in CATS for s in STATS]
)

BOROUGHS = ["M","BK","Q","BX","SI","FERRY"]
BOROUGH_FULL = {
    "M": "Manhattan",
    "BK": "Brooklyn",
    "Q": "Queens",
    "BX": "Bronx",
    "SI": "Staten Island",
    "FERRY": "Ferry"
}

ENV_FULL = {
    "C": "Commercial",
    "R": "Residential",
    "G": "Greenery",
    "T": "Transportation",
    "I": "Indoors"
}

# ---------------- HELPERS ----------------

def parse_filename(path):
    """Parse filename 'YYYY-MM-DD_hh.mm.ss E B location'."""
    name = os.path.splitext(os.path.basename(path))[0]
    m = re.match(
        r"(\d{4}-\d{2}-\d{2})(?:[_T]\d{2}\.\d{2}\.\d{2})?\s+([CRGTI])\s+(M|BK|Q|BX|SI|FERRY)\s+(.+)",
        name
    )
    if not m:
        return "", "", "", name
    date_str, env_type, borough, location = m.groups()
    return date_str, env_type, borough, location.strip()

def to_time_string(v):
    if isinstance(v, datetime): return v.strftime("%H:%M:%S")
    if isinstance(v, dtime):    return v.strftime("%H:%M:%S")
    if isinstance(v,str):
        for fmt in ("%m/%d/%Y %H:%M:%S","%m/%d/%Y %H:%M","%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(v.strip(),fmt).strftime("%H:%M:%S")
            except: pass
    return ""

def rss(df, cols):
    c = [x for x in cols if x in df.columns]
    if not c: return pd.Series(np.nan,index=df.index)
    return np.sqrt((df[c].apply(pd.to_numeric,errors="coerce")**2).sum(axis=1,min_count=1))

def geomean(x):
    x = np.asarray(x,dtype=float)
    x = x[x>0]
    return float(exp(np.mean(np.log(x)))) if x.size else np.nan

def stats(s):
    s = pd.Series(pd.to_numeric(s,errors="coerce")).dropna()
    if s.empty: return {k:np.nan for k in STATS}
    return {
        "MIN":s.min(),"P25":s.quantile(.25),"MEAN":s.mean(),"GEOMEAN":geomean(s),
        "MEDIAN":s.median(),"P75":s.quantile(.75),"P90":s.quantile(.9),
        "MAX":s.max(),"STDEV":s.std(ddof=1)
    }

def first_last_gps(df):
    """
    Find first and last rows where both 'GPS Lat' and 'GPS Lon' are valid numbers.
    Treat (0.0, 0.0) as invalid. Returns (start_lat, start_lon, end_lat, end_lon)
    as floats or np.nan if not found.
    """
    if "GPS Lat" not in df.columns or "GPS Lon" not in df.columns:
        return np.nan, np.nan, np.nan, np.nan

    lat_raw = df["GPS Lat"]
    lon_raw = df["GPS Lon"]

    lat_num = pd.to_numeric(lat_raw, errors="coerce")
    lon_num = pd.to_numeric(lon_raw, errors="coerce")

    # Valid if both numeric and not the (0,0) dummy position
    valid = lat_num.notna() & lon_num.notna()
    valid &= ~((lat_num == 0) & (lon_num == 0))

    if not valid.any():
        return np.nan, np.nan, np.nan, np.nan

    # First valid pair
    first_idx = valid[valid].index[0]
    start_lat = float(lat_num.loc[first_idx])
    start_lon = float(lon_num.loc[first_idx])

    # Last valid pair
    last_idx = valid[valid].index[-1]
    end_lat = float(lat_num.loc[last_idx])
    end_lon = float(lon_num.loc[last_idx])

    return start_lat, start_lon, end_lat, end_lon

def find_turn_gps(df):
    """
    Find the first row where Marker == 'x' (case-insensitive) and both
    'GPS Lat' and 'GPS Lon' are valid numbers and not (0.0, 0.0).
    Returns (turn_lat, turn_lon) as floats or np.nan if not found.
    """
    if "Marker" not in df.columns:
        return np.nan, np.nan
    if "GPS Lat" not in df.columns or "GPS Lon" not in df.columns:
        return np.nan, np.nan

    marker = df["Marker"].astype(str).str.strip().str.lower()
    mask_marker = marker == "x"

    if not mask_marker.any():
        return np.nan, np.nan

    lat_raw = df.loc[mask_marker, "GPS Lat"]
    lon_raw = df.loc[mask_marker, "GPS Lon"]

    lat_num = pd.to_numeric(lat_raw, errors="coerce")
    lon_num = pd.to_numeric(lon_raw, errors="coerce")

    valid = lat_num.notna() & lon_num.notna()
    valid &= ~((lat_num == 0) & (lon_num == 0))

    if not valid.any():
        return np.nan, np.nan

    idx = valid[valid].index[0]
    turn_lat = float(lat_num.loc[idx])
    turn_lon = float(lon_num.loc[idx])
    return turn_lat, turn_lon

# ---------- INVENTORY FILE HANDLING ----------

def open_inventory(inv_path, header):
    """Open or create workbook; explicitly write header row 1."""
    if os.path.exists(inv_path):
        wb = load_workbook(inv_path)
        ws = wb.active
        # Clean possible empty leading rows
        while ws.max_row > 1:
            row1 = [c.value for c in ws[1]]
            if not any(row1):
                ws.delete_rows(1)
            else:
                break
        # Rewrite header to ensure it matches INVENTORY_HEADER
        for i,val in enumerate(header,start=1):
            ws.cell(row=1,column=i,value=val)
    else:
        wb = Workbook()
        ws = wb.active
        if ws.max_row>0: ws.delete_rows(1,ws.max_row)
        for i,val in enumerate(header,start=1):
            ws.cell(row=1,column=i,value=val)
    return wb,ws

def append_row(ws, vals, header):
    if len(vals)<len(header):
        vals += [None]*(len(header)-len(vals))
    ws.append(vals)

# ---------- PROCESS SOURCE ----------

def process_file(path, agg_store):
    df = pd.read_excel(path)

    # Time information
    tcol = df.columns[0]
    times = df[tcol].dropna()
    N = len(times)
    st = to_time_string(times.iloc[0]) if N else ""
    et = to_time_string(times.iloc[-1]) if N else ""

    # GPS information: start/end
    start_lat, start_lon, end_lat, end_lon = first_last_gps(df)
    # GPS information: turn point (Marker == 'x')
    turn_lat, turn_lon = find_turn_gps(df)

    date, env, bor, loc = parse_filename(path)
    out = {
        "date": date,
        "borough": bor,
        "environment": env,
        "location": loc,
        "start": st,
        "end": et,
        "start_lat": None if np.isnan(start_lat) else start_lat,
        "start_lon": None if np.isnan(start_lon) else start_lon,
        "turn_lat": None if np.isnan(turn_lat) else turn_lat,
        "turn_lon": None if np.isnan(turn_lon) else turn_lon,
        "end_lat": None if np.isnan(end_lat) else end_lat,
        "end_lon": None if np.isnan(end_lon) else end_lon,
        "N": N,
        "stats": {}
    }

    for c in CATS:
        rss_vals = rss(df, CATEGORY_MAP[c]).dropna().astype(float)
        out["stats"][c] = stats(rss_vals)
        for label in ("Totals", bor, env):
            if label not in agg_store:
                agg_store[label] = {cat: [] for cat in CATS}
            agg_store[label][c].extend(rss_vals.tolist())
    return out

def build_row(r):
    # Order must match INVENTORY_HEADER
    vals = [
        r["date"],
        r["borough"],
        r["location"],
        r["environment"],
        "",                 # note
        r["start"],
        r["end"],
        r["start_lat"],
        r["start_lon"],
        r["turn_lat"],
        r["turn_lon"],
        r["end_lat"],
        r["end_lon"],
        r["N"],
    ]
    for c in CATS:
        for s in STATS:
            v = r["stats"][c][s]
            vals.append(None if np.isnan(v) else v)
    return vals

# ---------- AGGREGATED STATS (Sheet2) ----------

def write_table(ws, label, cat_dict):
    """Write one labeled statistics table with a blank line after it."""
    for cat in CATS:
        vals = cat_dict.get(cat, [])
        st = stats(vals)
        ws.append([label,cat]+[st[s] for s in STATS])
    ws.append([])  # single blank line between tables

def write_aggregates_sheet(wb, agg_store):
    """Write Sheet2 with formatted tables."""
    if "Sheet2" in wb.sheetnames:
        wb.remove(wb["Sheet2"])
    ws2 = wb.create_sheet("Sheet2")

    ws2.append(["Label","Category"]+STATS)

    # Order: Totals → Boroughs → Environments
    if "Totals" in agg_store:
        write_table(ws2, "Totals", agg_store["Totals"])

    for b in BOROUGHS:
        if b in agg_store:
            write_table(ws2, BOROUGH_FULL[b], agg_store[b])

    for e in ENV_FULL:
        if e in agg_store:
            write_table(ws2, ENV_FULL[e], agg_store[e])

    # ---------- Formatting ----------
    bold_font = Font(bold=True)

    # Bold first row
    for cell in ws2[1]:
        cell.font = bold_font

    # Bold first column
    for row in ws2.iter_rows(min_row=2):
        row[0].font = bold_font

    # Format all numeric cells to 4 decimal places
    for row in ws2.iter_rows(min_row=2, min_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0000"

    return ws2

# ---------- MAIN ----------

def main():
    Tk().withdraw()
    inv = filedialog.askopenfilename(title="Select INVENTORY Excel file",
                                     filetypes=[("Excel","*.xlsx *.xlsm")])
    if not inv: return
    srcs = filedialog.askopenfilenames(title="Select SOURCE Excel files",
                                       filetypes=[("Excel","*.xlsx *.xlsm")])
    if not srcs: return

    wb, ws = open_inventory(inv, INVENTORY_HEADER)
    agg_store = {}

    for s in srcs:
        try:
            r = process_file(s, agg_store)
            append_row(ws, build_row(r), INVENTORY_HEADER)
            print("Added", os.path.basename(s))
        except Exception as e:
            print("⚠️", os.path.basename(s), e)

    write_aggregates_sheet(wb, agg_store)

    wb.save(inv)
    print("\n✅ Inventory updated with start/turn/end GPS coordinates and formatted Totals, Borough, and Environment summaries in Sheet2:", inv)

if __name__=="__main__":
    main()
