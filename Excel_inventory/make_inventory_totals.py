"""
URBAN RF-EMF STUDY IN NYC
Source files created from ExpoM RF-EMF sensor source files with script csv2excel_batch.py
The filenames were manualy modified to include borough and location

Inventory Aggregation and Summary Script
----------------------------------------

This script automates the processing and aggregation of measurement data
collected from multiple Excel source files containing broadband RMS readings
across several frequency categories (Broadcast, Downlink, Uplink, WLAN, TDD, Total).

Main workflow:
1. Interactively prompts the user to:
   • Select an existing inventory Excel file (to be updated).
   • Select one or more source Excel files to process.

2. For each source file:
   • Parses the filename to extract:
       - date (YYYY-MM-DD or YYYY-MM-DD_hh.mm.ss → date only),
       - borough (one of [M, Q, BK, BX, SI, FERRY]),
       - location (remaining text after borough).
   • Reads the first column to determine:
       - start time (first timestamp),
       - end time (last timestamp),
       - N = number of measurement samples.
   • Computes Root-Sum-of-Squares (RSS) values per data row for each
     measurement category using the corresponding column group defined in CATEGORY_MAP.
   • Derives nine descriptive statistics for each category:
       [MIN, P25, MEAN, GEOMEAN, MEDIAN, P75, P90, MAX, STDEV].
   • Appends a single summary row to the inventory file (Sheet1) containing:
       [date, borough, location, type, note, start time, end time, N,
        Broadcast MIN, Broadcast P25, ..., Total STDEV].

3. While processing, the script also aggregates all raw RSS values into memory
   for later computation of *true* overall and per-borough statistics.

4. After all source files are processed:
   • Creates or replaces Sheet2 ("Aggregated Statistics") in the inventory file.
   • Computes the nine statistics directly from all combined raw RSS values:
       - One "Totals" section combining all files together.
       - One section per borough (M, Q, BK, BX, SI, FERRY) if data exist.
   • Writes the results as a table:
       Label | Category | MIN | P25 | MEAN | GEOMEAN | MEDIAN | P75 | P90 | MAX | STDEV

5. Saves the updated inventory file with both sheets:
   • Sheet1 - per-file summary (appended rows)
   • Sheet2 - aggregated “true” statistics based on all raw measurements.

Notes:
- The script ensures the header row in Sheet1 is always present exactly once (row 1).
- It can be run repeatedly: new data are appended, and Sheet2 is regenerated.
- Requires: pandas, numpy, openpyxl (install via `pip install pandas numpy openpyxl`).

Author: (RToledo-Crow + ChatGPT, ASRC CUNY)
Date: (October 2025)
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os, re
from math import exp, log
from datetime import datetime, time as dtime
import numpy as np, pandas as pd
from tkinter import Tk, filedialog
from openpyxl import load_workbook, Workbook

# ---------------- CONFIGURATION ----------------

CATEGORY_MAP = {
    "Broadcast": ["FM Radio (RMS)", "VHF 1, 2, 3 (RMS)", "UHF1 (RMS)", "UHF2 (RMS)", "UHF3 (RMS)"],
    "Downlink":  ["Mobile DL (RMS)", "Mobile DL (RMS).1", "Mobile DL (RMS).2", "Mobile DL (RMS).3", "Mobile DL (RMS).4"],
    "Uplink":    ["Mobile UL (RMS).1","Mobile UL (RMS).2","Mobile UL (RMS).3","Mobile UL (RMS).4","Mobile UL (RMS).5"],
    "WLAN":      ["ISM (RMS)","WLAN (RMS)","WLAN (RMS).1","WLAN (RMS).2","WLAN (RMS).3","WLAN (RMS).4",
                  "WLAN (RMS).5","WLAN (RMS).6","WLAN (RMS).7","WLAN (RMS).8","WLAN (RMS).9","WLAN (RMS).10"],
    "TDD":       ["TDD (RMS)","TDD (RMS).1","TDD (RMS).2","TDD (RMS).3","TDD (RMS).4",
                  "TDD (RMS).5","TDD (RMS).6","TDD (RMS).7","TDD (RMS).8"],
}
CATEGORY_MAP["Total"] = sum(CATEGORY_MAP.values(), [])

CATS = ["Broadcast","Downlink","Uplink","WLAN","TDD","Total"]
STATS = ["MIN","P25","MEAN","GEOMEAN","MEDIAN","P75","P90","MAX","STDEV"]
INVENTORY_HEADER = (
    ["date","borough","location","type","note","start time","end time","N"]
    + [f"{c} {s}" for c in CATS for s in STATS]
)
BOROUGHS = ["M","Q","BK","BX","SI","FERRY"]

# ---------------- HELPERS ----------------

def parse_filename(path):
    """Extract YYYY-MM-DD (ignore _hh.mm.ss), borough, location"""
    name = os.path.splitext(os.path.basename(path))[0]
    m = re.search(r"(\d{4}-\d{2}-\d{2})(?:[_T]\d{2}\.\d{2}\.\d{2})?", name)
    date_str = m.group(1) if m else ""
    parts = re.split(r"\s+", name)
    borough, loc = "", ""
    for i,p in enumerate(parts):
        if p in BOROUGHS:
            borough = p
            loc = " ".join(parts[i+1:])
            break
    return date_str, borough, loc.strip()

def to_time_string(v):
    if isinstance(v, datetime): return v.strftime("%H:%M:%S")
    if isinstance(v, dtime):    return v.strftime("%H:%M:%S")
    if isinstance(v,str):
        for fmt in ("%m/%d/%Y %H:%M:%S","%m/%d/%Y %H:%M","%Y-%m-%d %H:%M:%S"):
            try: return datetime.strptime(v.strip(),fmt).strftime("%H:%M:%S")
            except: pass
    return ""

def rss(df, cols):
    c = [x for x in cols if x in df.columns]
    if not c: return pd.Series(np.nan,index=df.index)
    return np.sqrt((df[c].apply(pd.to_numeric,errors="coerce")**2).sum(axis=1,min_count=1))

def geomean(x):
    x = np.asarray(x,dtype=float)
    x = x[x>0]
    return float(exp(np.mean(np.log(x)))) if x.size else np.nan

def stats(s):
    s = pd.Series(pd.to_numeric(s,errors="coerce")).dropna()
    if s.empty: return {k:np.nan for k in STATS}
    return {
        "MIN":s.min(),"P25":s.quantile(.25),"MEAN":s.mean(),"GEOMEAN":geomean(s),
        "MEDIAN":s.median(),"P75":s.quantile(.75),"P90":s.quantile(.9),
        "MAX":s.max(),"STDEV":s.std(ddof=1)
    }

# ---------- INVENTORY FILE HANDLING ----------

def open_inventory(inv_path, header):
    """Open or create workbook; explicitly write header row 1."""
    if os.path.exists(inv_path):
        wb = load_workbook(inv_path)
        ws = wb.active
        while ws.max_row > 1:
            row1 = [c.value for c in ws[1]]
            if not any(row1):
                ws.delete_rows(1)
            else:
                break
        for i,val in enumerate(header,start=1):
            ws.cell(row=1,column=i,value=val)
    else:
        wb = Workbook()
        ws = wb.active
        if ws.max_row>0: ws.delete_rows(1,ws.max_row)
        for i,val in enumerate(header,start=1):
            ws.cell(row=1,column=i,value=val)
    return wb,ws

def append_row(ws, vals, header):
    if len(vals)<len(header):
        vals += [None]*(len(header)-len(vals))
    ws.append(vals)

# ---------- PROCESS SOURCE ----------

def process_file(path, agg_store):
    df = pd.read_excel(path)
    tcol = df.columns[0]
    times = df[tcol].dropna()
    N=len(times)
    st=to_time_string(times.iloc[0]) if N else ""
    et=to_time_string(times.iloc[-1]) if N else ""
    date,bor,loc=parse_filename(path)
    out={"date":date,"borough":bor,"location":loc,"start":st,"end":et,"N":N,"stats":{}}

    # Compute per-category RSS vectors and add them to aggregate storage
    for c in CATS:
        rss_vals = rss(df, CATEGORY_MAP[c]).dropna().astype(float)
        out["stats"][c]=stats(rss_vals)
        # store values globally
        if "Totals" not in agg_store:
            agg_store["Totals"] = {cat: [] for cat in CATS}
        if bor not in agg_store:
            agg_store[bor] = {cat: [] for cat in CATS}
        agg_store["Totals"][c].extend(rss_vals.tolist())
        agg_store[bor][c].extend(rss_vals.tolist())

    return out

def build_row(r):
    vals=[r["date"],r["borough"],r["location"],"", "",r["start"],r["end"],r["N"]]
    for c in CATS:
        for s in STATS:
            v=r["stats"][c][s]
            vals.append(None if np.isnan(v) else v)
    return vals

# ---------- AGGREGATED STATS (Sheet2) ----------

def write_aggregates_sheet(wb, agg_store):
    """Create Sheet2 with stats computed on total raw values."""
    if "Sheet2" in wb.sheetnames:
        wb.remove(wb["Sheet2"])
    ws2 = wb.create_sheet("Sheet2")

    ws2.append(["Label","Category"]+STATS)
    for label, cat_dict in agg_store.items():
        for cat in CATS:
            vals = cat_dict.get(cat, [])
            st = stats(vals)
            ws2.append([label,cat]+[st[s] for s in STATS])
        ws2.append([])

    return ws2

# ---------- MAIN ----------

def main():
    Tk().withdraw()
    inv = filedialog.askopenfilename(title="Select INVENTORY Excel file",
                                     filetypes=[("Excel","*.xlsx *.xlsm")])
    if not inv: return
    srcs = filedialog.askopenfilenames(title="Select SOURCE Excel files",
                                       filetypes=[("Excel","*.xlsx *.xlsm")])
    if not srcs: return

    wb, ws = open_inventory(inv, INVENTORY_HEADER)
    agg_store = {}  # dictionary to hold all raw RSS values per category

    for s in srcs:
        try:
            r=process_file(s, agg_store)
            append_row(ws,build_row(r),INVENTORY_HEADER)
            print("Added",os.path.basename(s))
        except Exception as e:
            print("⚠️",os.path.basename(s),e)

    # ---- Aggregated raw-data statistics to Sheet2 ----
    write_aggregates_sheet(wb, agg_store)

    wb.save(inv)
    print("\n✅ Inventory updated with true aggregated statistics (Sheet2):", inv)

if __name__=="__main__":
    main()
